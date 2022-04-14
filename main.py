from flask import Flask, render_template, request, redirect, url_for
import openpyxl

path = "Sito.xlsx"
wb_obj = openpyxl.load_workbook(path) 
sheet_obj = wb_obj.active
wb = openpyxl.Workbook() 
sheet = wb.active

righe = sheet_obj.max_row
colonne = sheet_obj.max_column

listaemail=[]
for i in range(1, righe + 1): 
  cell_obj = sheet_obj.cell(row = i, column = 1) 
  listaemail.append(cell_obj)

app = Flask(
	__name__,
	template_folder='templates',
	static_folder='static'  
)

@app.route('/')
def intro_page():
	return render_template('intro.html')

@app.route('/login')
def login():
	return render_template('login.html')

@app.route('/login',methods = ['POST', 'GET'])
def loginerrori():
  LoginPassword = request.form["Password"]
  LoginEmail = request.form["Email"]
  if LoginPassword == "":
    return render_template('errore2.html')
  elif LoginEmail == "":
    return render_template('errore2.html')
  else:
    return redirect(url_for('base_page'))

@app.route('/register')
def register():
  return render_template('register.html')

@app.route('/register',methods = ['POST', 'GET'])
def errori():
  Password = request.form["Password"]
  Email = request.form["Email"]
  Nome = request.form["Nome"]
  Cognome = request.form["Cognome"]
  RipetiPassword = request.form["RipetiPassword"]
  if Password == "":
    return render_template('errore2.html')
  elif RipetiPassword == "":
    return render_template('errore2.html')
  elif Nome == "":
    return render_template('errore2.html')
  elif Cognome == "":
    return render_template('errore2.html')
  elif Email == "":
    return render_template('errore2.html')
  elif Password!=RipetiPassword:
    return render_template('errore1.html')
  if Email in listaemail:
    return render_template('errore3.html')
  else:
    cell=sheet["A"+str(righe+1)]
    cell.value=Email
    wb.save("Sito.xlsx")
    return redirect(url_for('base_page'))
  
@app.route('/shop')
def base_page():
  return render_template('base.html')

@app.route('/contact-us')
def contact_us():
	return render_template('contact_us.html')

@app.route('/about-us')
def about_us():
	return render_template('about_us.html')

@app.route('/aleks')
def aleks():
	return render_template('aleks.html')

if __name__ == "__main__":
  app.run(
		host='0.0.0.0',
		port=6969,
    debug=True
	)
